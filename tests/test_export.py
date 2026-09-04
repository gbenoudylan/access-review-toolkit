"""Test du module d'export (Excel + PDF) pour la revue d'accès."""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import pandas as pd
from reporting.export import generate_excel_report, generate_pdf_report


def build_sample_analyzed_df() -> pd.DataFrame:
    return pd.DataFrame({
        "username": ["jdupont", "kbrou", "mfofana", "sassoum"],
        "full_name": ["Jean Dupont", "Konan Brou", "Marc Fofana", "Sara Assoum"],
        "department": ["IT", "IT Security", "Sales", "HR"],
        "system": ["Active Directory", "SIEM", "CRM", "HRIS"],
        "manager": ["Marie D.", "", "Marie D.", "Paul N."],
        "account_status": ["Active", "Active", "Active", "Active"],
        "employee_status": ["Active", "Active", "Terminated", "Active"],
        "days_since_last_login": [2, 210, 5, 1],
        "is_privileged_flag": [False, True, False, False],
        "is_terminated_but_active": [False, False, True, False],
        "is_dormant": [False, True, False, False],
        "review_action": [
            "Aucune action", "Désactiver (privilégié dormant)",
            "Révoquer immédiatement", "Aucune action",
        ],
        "risk_level": ["Faible", "Critique", "Critique", "Faible"],
    })


def test_generate_excel_report(tmp_path):
    df = build_sample_analyzed_df()
    output = tmp_path / "test.xlsx"
    result = generate_excel_report(df, output)
    assert result.exists() and result.stat().st_size > 0

    from openpyxl import load_workbook
    wb = load_workbook(result)
    assert "Synthèse" in wb.sheetnames
    assert "Plan de revue" in wb.sheetnames
    print(f"OK - test_generate_excel_report ({result.stat().st_size} octets)")


def test_generate_pdf_report(tmp_path):
    df = build_sample_analyzed_df()
    output = tmp_path / "test.pdf"
    result = generate_pdf_report(df, output)
    assert result.exists()
    with open(result, "rb") as f:
        assert f.read(5) == b"%PDF-"
    print(f"OK - test_generate_pdf_report ({result.stat().st_size} octets)")


if __name__ == "__main__":
    import tempfile
    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_path = Path(tmp_dir)
        test_generate_excel_report(tmp_path)
        test_generate_pdf_report(tmp_path)
    print("\nTous les tests sont passés.")


def test_pdf_report_handles_missing_values_without_crashing():
    """
    .astype(str) sur un DataFrame ne convertit pas les valeurs manquantes
    (NaN) en texte — elles restent des float et font planter Paragraph()
    dans le tableau PDF, qui exige une vraie chaîne. Ce test couvre
    explicitement des comptes avec système ou nom manquant.
    """
    import pandas as pd
    from analysis.access_review import analyze_access
    from reporting.export import generate_pdf_report

    df = pd.DataFrame({
        "username": ["jdupont", "kbrou", "mfofana"],
        "full_name": ["Jean Dupont", None, "Marc Fofana"],
        "system": ["Active Directory", "CRM", None],
        "manager": [None, "Paul N.", "Marie D."],
    })
    result = analyze_access(df)
    output = generate_pdf_report(result, "output/test_missing_values.pdf")
    assert output.exists()
    print("OK - test_pdf_report_handles_missing_values_without_crashing")


def test_pdf_report_includes_signoff_names_when_provided():
    """Les noms de validation fournis doivent apparaître dans le PDF généré."""
    import pandas as pd
    from analysis.access_review import analyze_access
    from reporting.export import generate_pdf_report
    import pdfplumber

    df = pd.DataFrame({"username": ["jdupont"], "system": ["Active Directory"]})
    result = analyze_access(df)
    output = generate_pdf_report(
        result, "output/test_signoff.pdf",
        prepared_by="Test Preparateur", reviewed_by="Test Revu", approved_by="Test Approuve",
    )
    with pdfplumber.open(output) as pdf:
        full_text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    assert "Test Preparateur" in full_text
    assert "Test Revu" in full_text
    assert "Test Approuve" in full_text
    assert "Objectifs" in full_text
    assert "Procédure" in full_text
    assert "Périmètre" in full_text
    print("OK - test_pdf_report_includes_signoff_names_when_provided")
