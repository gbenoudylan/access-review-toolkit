"""
Module d'export du rapport de revue d'accès.

Génère deux livrables à partir du DataFrame analysé (sortie de
analysis/access_review.py) :

    - Un rapport Excel détaillé avec mise en forme conditionnelle par
      niveau de risque, prêt pour le suivi opérationnel.
    - Un rapport PDF de synthèse, adapté à une diffusion managériale ou
      une preuve d'audit pour la revue d'accès périodique.
"""

from __future__ import annotations
import logging
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

logger = logging.getLogger("export")

RISK_COLORS_HEX = {
    "Critique": "D62728",
    "Élevé": "FF7F0E",
    "Moyen": "FFD700",
    "Faible": "2CA02C",
}

# Formulation générique associée à chaque action recommandée, pour la
# section narrative "Rapport des exceptions" — inspirée des standards du
# secteur (revue trimestrielle des accès), jamais copiée d'un document
# précis : ces recommandations sont volontairement génériques pour rester
# valables quelle que soit l'entreprise ou le système concerné.
ACTION_NARRATIVE = {
    "Révoquer immédiatement": (
        "Ces comptes restent actifs alors que la personne associée a quitté "
        "l'entreprise. Action recommandée : révocation immédiate des accès."
    ),
    "Désactiver (privilégié dormant)": (
        "Ces comptes disposent de privilèges élevés et n'ont enregistré aucune "
        "connexion depuis le seuil de dormance retenu. Action recommandée : "
        "désactivation, le niveau d'accès concerné justifie une vigilance "
        "renforcée."
    ),
    "Forcer l'expiration du mot de passe (privilégié)": (
        "Ces comptes à privilèges élevés ont un mot de passe configuré pour "
        "ne jamais expirer. Action recommandée : appliquer une politique "
        "d'expiration standard, et documenter toute exception justifiée "
        "(compte de service avec surveillance dédiée)."
    ),
    "Désactiver (dormant)": (
        "Ces comptes n'ont enregistré aucune connexion depuis le seuil de "
        "dormance retenu. Action recommandée : vérifier auprès du "
        "propriétaire métier, puis désactiver si l'usage n'est plus justifié."
    ),
    "Exiger un changement de mot de passe": (
        "Le mot de passe de ces comptes n'a pas été renouvelé depuis le seuil "
        "retenu. Action recommandée : forcer le changement à la prochaine "
        "connexion."
    ),
    "Identifier un owner": (
        "Aucun manager ou propriétaire métier n'est identifié pour ces "
        "comptes. Action recommandée : désigner un responsable chargé de "
        "valider la légitimité de l'accès."
    ),
    "Vérifier avec le propriétaire technique (compte de service)": (
        "Ces comptes de service n'ont enregistré aucune activité depuis le "
        "seuil de dormance retenu. Action recommandée : vérifier auprès du "
        "propriétaire technique s'ils sont toujours utilisés par un "
        "processus automatisé avant toute décision, une désactivation "
        "directe pouvant casser un traitement encore actif."
    ),
    "Fusionner les doublons (ne garder qu'un compte actif)": (
        "Plusieurs comptes actifs semblent appartenir à la même personne "
        "sur le même système. Action recommandée : confirmer le doublon "
        "auprès du titulaire, puis désactiver tous les comptes superflus "
        "pour n'en garder qu'un seul actif."
    ),
}

DISPLAY_COLUMNS = [
    ("username", "Compte"),
    ("user_id", "ID employé"),
    ("full_name", "Nom"),
    ("department", "Département"),
    ("system", "Système"),
    ("manager", "Manager"),
    ("account_status", "Statut compte"),
    ("employee_status", "Statut RH"),
    ("days_since_last_login", "Jours sans connexion"),
    ("days_since_password_change", "Jours sans changement MDP"),
    ("is_privileged_flag", "Privilégié"),
    ("has_non_expiring_password", "MDP n'expire jamais"),
    ("review_action", "Action recommandée"),
    ("risk_level", "Risque"),
]


def _prepare_export_df(df: pd.DataFrame) -> pd.DataFrame:
    available = [(col, label) for col, label in DISPLAY_COLUMNS if col in df.columns]
    export_df = df[[col for col, _ in available]].copy()
    export_df.columns = [label for _, label in available]

    risk_order = {"Critique": 0, "Élevé": 1, "Moyen": 2, "Faible": 3}
    if "Risque" in export_df.columns:
        export_df["_sort"] = export_df["Risque"].map(risk_order).fillna(99)
        export_df = export_df.sort_values("_sort").drop(columns="_sort")
    return export_df


# ---------------------------------------------------------------------
# EXCEL
# ---------------------------------------------------------------------

def generate_excel_report(df: pd.DataFrame, output_path: str | Path) -> Path:
    output_path = Path(output_path)
    export_df = _prepare_export_df(df)

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Synthèse"

    ws_summary["A1"] = "Rapport de revue d'accès"
    ws_summary["A1"].font = Font(size=14, bold=True)
    ws_summary["A2"] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws_summary["A2"].font = Font(italic=True, color="666666")

    ws_summary["A4"] = "Niveau de risque"
    ws_summary["B4"] = "Nombre de comptes"
    ws_summary["A4"].font = ws_summary["B4"].font = Font(bold=True)

    risk_counts = df["risk_level"].value_counts() if "risk_level" in df.columns else {}
    row = 5
    for risk, hex_color in RISK_COLORS_HEX.items():
        count = int(risk_counts.get(risk, 0))
        ws_summary[f"A{row}"] = risk
        ws_summary[f"B{row}"] = count
        ws_summary[f"A{row}"].fill = PatternFill("solid", fgColor=hex_color)
        ws_summary[f"A{row}"].font = Font(color="FFFFFF", bold=True)
        row += 1

    ws_summary[f"A{row + 1}"] = "Total comptes analysés"
    ws_summary[f"B{row + 1}"] = len(df)
    ws_summary[f"A{row + 1}"].font = Font(bold=True)

    if "is_terminated_but_active" in df.columns:
        ws_summary[f"A{row + 3}"] = "Comptes actifs d'employés partis"
        ws_summary[f"B{row + 3}"] = int(df["is_terminated_but_active"].sum())
    if "is_dormant" in df.columns:
        ws_summary[f"A{row + 4}"] = "Comptes dormants"
        ws_summary[f"B{row + 4}"] = int(df["is_dormant"].sum())
    if "is_password_stale" in df.columns:
        ws_summary[f"A{row + 5}"] = "Mots de passe périmés"
        ws_summary[f"B{row + 5}"] = int(df["is_password_stale"].sum())
    if "is_privileged_flag" in df.columns and "has_non_expiring_password" in df.columns:
        ws_summary[f"A{row + 6}"] = "Comptes privilégiés à mot de passe n'expirant jamais"
        ws_summary[f"B{row + 6}"] = int((df["is_privileged_flag"] & df["has_non_expiring_password"]).sum())
    if "is_duplicate_account" in df.columns:
        ws_summary[f"A{row + 7}"] = "Comptes en doublon"
        ws_summary[f"B{row + 7}"] = int(df["is_duplicate_account"].sum())

    for col, width in zip("AB", [32, 20]):
        ws_summary.column_dimensions[col].width = width

    ws = wb.create_sheet("Plan de revue")
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(*[Side(style="thin", color="D9D9D9")] * 4)

    for col_idx, col_name in enumerate(export_df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill, cell.font = header_fill, header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    risk_col_idx = (
        list(export_df.columns).index("Risque") + 1 if "Risque" in export_df.columns else None
    )

    for row_idx, record in enumerate(export_df.to_dict("records"), 2):
        for col_idx, (col_name, value) in enumerate(record.items(), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
        if risk_col_idx:
            hex_color = RISK_COLORS_HEX.get(record.get("Risque"))
            if hex_color:
                cell = ws.cell(row=row_idx, column=risk_col_idx)
                cell.fill = PatternFill("solid", fgColor=hex_color)
                cell.font = Font(color="FFFFFF", bold=True)

    for col_idx, col_name in enumerate(export_df.columns, 1):
        max_len = max([len(str(col_name))] + [len(str(v)) for v in export_df[col_name].astype(str)])
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info(f"Rapport Excel généré : {output_path}")
    return output_path


# ---------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------

def _current_quarter_label() -> str:
    now = datetime.now()
    quarter = (now.month - 1) // 3 + 1
    return f"T{quarter} {now.year}"


# Poids relatifs de largeur par colonne (les colonnes non listées ont un
# poids par défaut de 1.0). "Action recommandée" et "Nom" sont plus larges
# car elles contiennent le texte le plus long — sans ça, ReportLab
# dimensionne les colonnes selon leur seul contenu, sans jamais tenir
# compte de la largeur réelle de la page, d'où un tableau qui déborde.
COLUMN_WIDTH_WEIGHTS = {
    "Compte": 1.1,
    "Nom": 1.4,
    "Département": 1.0,
    "Système": 1.0,
    "Manager": 1.0,
    "Statut compte": 0.9,
    "Statut RH": 0.9,
    "Jours sans connexion": 0.9,
    "Jours sans changement MDP": 1.1,
    "Privilégié": 0.7,
    "MDP n'expire jamais": 1.0,
    "Action recommandée": 2.2,
    "Risque": 0.8,
}
# Colonnes dont le texte doit pouvoir revenir à la ligne plutôt que
# déborder ou être tronqué.
WRAP_COLUMNS = {"Nom", "Système", "Action recommandée"}


def _compute_column_widths(columns: list[str], available_width: float) -> list[float]:
    weights = [COLUMN_WIDTH_WEIGHTS.get(col, 1.0) for col in columns]
    total_weight = sum(weights)
    return [available_width * w / total_weight for w in weights]


def _risk_styled_table(export_df: pd.DataFrame, available_width: float) -> Table:
    """Construit une table stylée (en-tête sombre, lignes alternées, cellule
    Risque colorée), avec des largeurs de colonnes proportionnelles à la
    largeur réelle de la page plutôt qu'au seul contenu, et un retour à la
    ligne automatique — sur les en-têtes ET sur les colonnes de texte long
    (sans quoi un libellé de colonne trop long déborde silencieusement sur
    la colonne voisine plutôt que de simplement passer à la ligne)."""
    cell_style = ParagraphStyle("Cell", fontSize=7.5, leading=9, fontName="Helvetica")
    header_style = ParagraphStyle(
        "CellHeader", fontSize=7.5, leading=9, fontName="Helvetica-Bold", textColor=colors.white,
    )
    columns = list(export_df.columns)
    col_widths = _compute_column_widths(columns, available_width)

    header_row = [Paragraph(str(col), header_style) for col in columns]
    data_rows = []
    # .astype(str) ne convertit pas les valeurs manquantes (NaN) en texte —
    # elles restent des float et font planter Paragraph() plus bas, qui
    # exige une vraie chaîne. On les remplace explicitement avant conversion.
    for record in export_df.fillna("").astype(str).values.tolist():
        row = []
        for col_name, value in zip(columns, record):
            if col_name in WRAP_COLUMNS:
                row.append(Paragraph(value, cell_style))
            else:
                row.append(value)
        data_rows.append(row)

    table = Table([header_row] + data_rows, repeatRows=1, colWidths=col_widths)

    style_commands = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7.5),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9F9F9")]),
    ]
    if "Risque" in export_df.columns:
        risk_col_idx = columns.index("Risque")
        for row_idx, risk_value in enumerate(export_df["Risque"], 1):
            hex_color = RISK_COLORS_HEX.get(risk_value)
            if hex_color:
                style_commands.append((
                    "BACKGROUND", (risk_col_idx, row_idx), (risk_col_idx, row_idx),
                    colors.HexColor(f"#{hex_color}"),
                ))
                style_commands.append((
                    "TEXTCOLOR", (risk_col_idx, row_idx), (risk_col_idx, row_idx), colors.white,
                ))
    table.setStyle(TableStyle(style_commands))
    return table


def _build_exceptions_section(df: pd.DataFrame, section_style, exception_style, action_style) -> list:
    """
    Construit la section narrative "Rapport des exceptions" : les comptes
    signalés sont regroupés par (système, action recommandée), puis
    numérotés "Exception N : ... Action : ...", au format d'un rapport
    d'audit classique — plutôt que le tableau brut de la section suivante.
    """
    elements = [Paragraph("Rapport des exceptions", section_style)]

    if "system" not in df.columns or "review_action" not in df.columns:
        elements.append(Paragraph(
            "Champs insuffisants pour générer le rapport des exceptions "
            "(système et action recommandée requis).",
            action_style,
        ))
        return elements

    flagged = df[df["review_action"] != "Aucune action"]
    if flagged.empty:
        elements.append(Paragraph("Aucune exception à signaler sur ce cycle.", action_style))
        return elements

    counter = 1
    for system_name, system_group in flagged.groupby("system"):
        for action, action_group in system_group.groupby("review_action"):
            count = len(action_group)
            elements.append(Paragraph(
                f"<b>Exception {counter} — {system_name} :</b> {count} compte(s) "
                f"avec le statut « {action} ».",
                exception_style,
            ))
            narrative = ACTION_NARRATIVE.get(
                action, "Action recommandée : voir le détail par système ci-dessous."
            )
            elements.append(Paragraph(narrative, action_style))
            counter += 1

    return elements


def generate_pdf_report(
    df: pd.DataFrame,
    output_path: str | Path,
    period: str | None = None,
    dormant_threshold_days: int = 90,
    prepared_by: str | None = None,
    reviewed_by: str | None = None,
    approved_by: str | None = None,
) -> Path:
    """
    Génère un rapport PDF de revue d'accès structuré et réutilisable d'un
    cycle à l'autre : résumé exécutif, méthodologie, actions prioritaires,
    puis détail système par système — plutôt qu'un unique tableau brut.

    `period` est un libellé libre (ex. "T1 2026", "Mars 2026") affiché en
    en-tête du rapport ; par défaut, le trimestre courant est déduit
    automatiquement, pour que la fonction reste utilisable telle quelle à
    chaque exécution sans argument supplémentaire.

    `prepared_by`, `reviewed_by`, `approved_by` : noms affichés dans le
    tableau de validation en fin de rapport. Laissés vides, ils affichent
    un espace à remplir à la main plutôt que de faire échouer la génération.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    period_label = period or _current_quarter_label()

    doc = SimpleDocTemplate(
        str(output_path), pagesize=landscape(A4),
        topMargin=1.5 * cm, bottomMargin=1.5 * cm, leftMargin=1.5 * cm, rightMargin=1.5 * cm,
    )
    available_width = doc.pagesize[0] - doc.leftMargin - doc.rightMargin
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TitleCustom", parent=styles["Title"], fontSize=18, spaceAfter=4)
    subtitle_style = ParagraphStyle("Subtitle", parent=styles["Normal"], fontSize=10, textColor=colors.grey)
    section_style = ParagraphStyle("SectionH", parent=styles["Heading2"], spaceBefore=14, spaceAfter=6)
    system_style = ParagraphStyle(
        "SystemH", parent=styles["Heading3"], textColor=colors.HexColor("#1F2937"),
        spaceBefore=12, spaceAfter=4,
    )
    note_style = ParagraphStyle("Note", parent=styles["Normal"], fontSize=8.5, textColor=colors.grey, spaceAfter=10)
    exception_style = ParagraphStyle(
        "Exception", parent=styles["Normal"], fontSize=9.5, spaceBefore=8, spaceAfter=2,
    )
    action_style = ParagraphStyle(
        "ActionText", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#374151"), spaceAfter=4,
    )

    elements = [
        Paragraph("Rapport de revue d'accès", title_style),
        Paragraph(f"Période : {period_label} — généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", subtitle_style),
        Spacer(1, 0.5 * cm),
    ]

    # ---- Objectifs (standards génériques d'une revue d'accès, valables
    # quel que soit le système ou l'entreprise concernée) ----
    elements.append(Paragraph("Objectifs", section_style))
    elements.append(Paragraph(
        "Cette revue vise à s'assurer que les comptes existants respectent les "
        "critères de sécurité attendus, notamment :",
        note_style,
    ))
    objectives = [
        "Absence de comptes non documentés ou non autorisés.",
        "Cohérence des privilèges accordés avec le besoin réel.",
        "Identification et traitement des comptes inactifs ou orphelins.",
        "Accès restreints de façon à limiter le risque de compromission.",
    ]
    for obj in objectives:
        elements.append(Paragraph(f"•&nbsp;&nbsp;{obj}", note_style))
    elements.append(Spacer(1, 0.2 * cm))

    # ---- Procédure (méthodologie standard, en 4 étapes) ----
    elements.append(Paragraph("Procédure", section_style))
    procedure_steps = [
        "Collecte de la liste des comptes et des journaux d'activité auprès des équipes concernées.",
        "Revue des accès et des événements au regard des objectifs ci-dessus.",
        "Constitution du rapport des exceptions.",
        "Validation du rapport et clôture des exceptions.",
    ]
    for i, step in enumerate(procedure_steps, 1):
        elements.append(Paragraph(f"{i}. {step}", note_style))
    elements.append(Spacer(1, 0.2 * cm))

    # ---- Scope (dérivé des données réelles, pas déclaratif) ----
    elements.append(Paragraph("Périmètre (Scope)", section_style))
    if "system" in df.columns:
        systems_in_scope = sorted(df["system"].dropna().unique().tolist())
        scope_text = ", ".join(systems_in_scope) if systems_in_scope else "Non renseigné"
    else:
        scope_text = "Non renseigné (colonne 'system' absente)"
    elements.append(Paragraph(f"Systèmes couverts par cette revue : {scope_text}.", note_style))
    elements.append(Spacer(1, 0.3 * cm))

    # ---- Méthodologie (courte, pour rappeler le seuil appliqué) ----
    elements.append(Paragraph(
        f"Méthodologie : un compte est considéré « dormant » sans connexion depuis plus de "
        f"{dormant_threshold_days} jours. Chaque compte reçoit un niveau de risque et une action "
        f"recommandée selon son statut (compte actif d'un employé parti, compte privilégié "
        f"dormant, absence de manager identifié).",
        note_style,
    ))

    # ---- Résumé exécutif ----
    elements.append(Paragraph("Résumé exécutif", section_style))
    risk_counts = df["risk_level"].value_counts() if "risk_level" in df.columns else {}
    summary_data = [["Indicateur", "Valeur"], ["Total comptes analysés", str(len(df))]]
    for risk in RISK_COLORS_HEX:
        summary_data.append([risk, str(int(risk_counts.get(risk, 0)))])
    if "is_terminated_but_active" in df.columns:
        summary_data.append(["Comptes actifs d'employés partis", str(int(df["is_terminated_but_active"].sum()))])
    if "is_dormant" in df.columns:
        summary_data.append(["Comptes dormants", str(int(df["is_dormant"].sum()))])
    if "is_password_stale" in df.columns:
        summary_data.append(["Mots de passe périmés", str(int(df["is_password_stale"].sum()))])
    if "is_privileged_flag" in df.columns and "has_non_expiring_password" in df.columns:
        summary_data.append([
            "Comptes privilégiés à mot de passe n'expirant jamais",
            str(int((df["is_privileged_flag"] & df["has_non_expiring_password"]).sum())),
        ])
    if "is_duplicate_account" in df.columns:
        summary_data.append(["Comptes en doublon", str(int(df["is_duplicate_account"].sum()))])

    summary_table = Table(summary_data, colWidths=[9 * cm, 4 * cm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
    ]))
    elements.append(summary_table)

    export_df_full = _prepare_export_df(df)

    # ---- Actions prioritaires (Critique + Élevé, tous systèmes confondus) ----
    if "Risque" in export_df_full.columns:
        priority_df = export_df_full[export_df_full["Risque"].isin(["Critique", "Élevé"])]
        elements.append(Paragraph(
            f"Actions prioritaires ({len(priority_df)} compte(s) à traiter en premier)",
            section_style,
        ))
        if len(priority_df):
            elements.append(_risk_styled_table(priority_df, available_width))
        else:
            elements.append(Paragraph("Aucun compte en risque Critique ou Élevé sur ce cycle.", styles["Normal"]))

    # ---- Rapport des exceptions (narratif, format audit classique) ----
    elements.extend(_build_exceptions_section(df, section_style, exception_style, action_style))

    # ---- Détail par système ----
    elements.append(Paragraph("Détail par système", section_style))
    if "system" in df.columns and "Système" in export_df_full.columns:
        systems = sorted(export_df_full["Système"].dropna().unique().tolist())
        for system_name in systems:
            system_df = export_df_full[export_df_full["Système"] == system_name]
            elements.append(Paragraph(f"{system_name} — {len(system_df)} compte(s)", system_style))
            elements.append(_risk_styled_table(system_df, available_width))
            elements.append(Spacer(1, 0.4 * cm))
    else:
        elements.append(Paragraph("Détail des comptes (triés par risque)", system_style))
        elements.append(_risk_styled_table(export_df_full, available_width))

    # ---- Sign-off (validation) ----
    elements.append(Paragraph("Validation", section_style))
    placeholder = "[À compléter]"
    signoff_data = [
        ["Rôle", "Nom", "Date"],
        ["Préparé par", prepared_by or placeholder, datetime.now().strftime("%d/%m/%Y")],
        ["Revu par", reviewed_by or placeholder, ""],
        ["Approuvé par", approved_by or placeholder, ""],
    ]
    signoff_table = Table(signoff_data, colWidths=[5 * cm, 8 * cm, 4 * cm])
    signoff_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(signoff_table)

    doc.build(elements)
    logger.info(f"Rapport PDF généré ({period_label}) : {output_path}")
    return output_path


if __name__ == "__main__":
    import sys
    from pathlib import Path as _Path
    sys.path.insert(0, str(_Path(__file__).parent.parent))

    logging.basicConfig(level=logging.INFO, format="%(levelname)s | %(message)s")

    from ingestion.ingest import load_file
    from analysis.access_review import analyze_access

    if len(sys.argv) < 2:
        print("Usage : python -m reporting.export <chemin_fichier>")
        sys.exit(1)

    df_result = analyze_access(load_file(sys.argv[1]))
    out_dir = _Path(__file__).parent.parent / "output"
    generate_excel_report(df_result, out_dir / "rapport_revue_acces.xlsx")
    generate_pdf_report(df_result, out_dir / "rapport_revue_acces.pdf")
